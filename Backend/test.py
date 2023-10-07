from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import openpyxl

# Variables 
lara_url = "https://cofs.lara.state.mi.us/SearchApi/Search/Search"
owner1 = "G.M.K., L.L.C."
owner2 = "AB11 CORPORATION"

# Initialize excel
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Create a set to store unique names
unique_names = set()

# LARA
def startLara(url, owner):
    # Find chromedriver on machine
    s = Service('../chromedriver/chromedriver')
    driver = webdriver.Chrome(service = s)

    # Open url in chrome
    driver.get(url)
    time.sleep(3)

    driver.find_element(By.ID, "txtEntityName").send_keys(owner)
    time.sleep(3)

    driver.find_element(By.ID, "SearchSubmit").click()
    time.sleep(3)

    # TODO IF THERE ARE MULPITPLE OPTIONS FOR LLC CHECK THE ADDRESS FROM LANDVISON
    driver.find_element(By.XPATH, f"//a[contains(text(), '{owner}')]").click()
    time.sleep(3)

    resident_element = driver.find_element(By.ID, "MainContent_lblResidentAgentName")
    resident = resident_element.text
    names = resident.split()

    if len(names) >= 2:
        first_name = names[0]
        last_name = names[1]

        # Check if the name is not already in the set
        if (first_name, last_name) not in unique_names:
            # Add the name to the set of unique names
            unique_names.add((first_name, last_name))

            # Write names to the Excel file
            row = len(unique_names)  # Get the row number for this name
            worksheet.cell(row=row, column=1, value=first_name)
            worksheet.cell(row=row, column=2, value=last_name)

            # Save the workbook and close the driver
            workbook.save('../excel/output_test.xlsx')
            driver.quit()

        # Return the first and last name (optional)
        return first_name
        return last_name
    else: 
        print("The text does not contain two words.")


# Method calls
startLara(lara_url, owner1)
startLara(lara_url, owner2)